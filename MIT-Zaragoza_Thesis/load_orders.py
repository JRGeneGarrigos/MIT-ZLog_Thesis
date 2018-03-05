# -*- coding: utf-8 -*-
"""
Created on Fri Feb 23 00:49:06 2018

@author: JRGene
"""

from routes_graph import SmartOrder, Transport
import load_network as ln
import openpyxl
import datetime
import pandas as pd
import os
import string

def load_Orders(Filename, graph, hash_nodes):
    """
    Parses the map file and constructs a directed graph

    Parameters:
        mapFilename : name of the map file

    Assumes:
        One line per each edge-entry in the map file consists of two strings 
        (origin and destiny) and one (integer leadtime), all separated by a 
        blank space:
            From To TotalLeadtime
        e.g.
            Houston Rotterdam 1 7
            Rotterdam Singapore 2 7
        This entry would become an edge from Houston to Rotterdam.

    Returns:
        It is going to return schedule_link_reduced
    """
    print("Loading Orders map from file ", Filename)
    t = Transport()
    lines = open(Filename, 'r').readlines()
    names = 0 # name introduce temporally
    last_arrival = datetime.datetime(1995, 10, 2, 0, 0, 0, 0)
    for i in lines:
        names += 1 # name introduce temporally
        line = i.split()
        id_order = int(line[0])
        start_node = graph.getNode(str(hash_nodes[line[1]]))
        end_node = graph.getNode(str(hash_nodes[line[2]]))
        date_start = datetime.datetime.strptime(line[3],'%d/%m/%Y')
        date_end = datetime.datetime.strptime(line[4],'%d/%m/%Y')
        weight = float(line[5])
        name = 'Order ' + str(names) # name introduce temporally
        order = SmartOrder( id_order, start_node, end_node, \
                           date_start, date_end, weight, name)
        t.addOrder(order)
        t_order = list(t.orders)
        
        all_cand_routes = []
        for j in t_order:
            container = []
            [container.append(k) for k in ln.shortestPath(graph, j.getSource(), \
                              j.getDestination()) if k not in container]
            all_cand_routes.append(container)
        if last_arrival < date_end:
            last_arrival = date_end

    print (all_cand_routes)
    
    all_edges = []
    [all_edges.append(graph.getEdgeDestination(j[k],j[k+1]))\
     for i in all_cand_routes for j in i for k in range(len(j)-1)\
     if graph.getEdgeDestination(j[k],j[k+1]) not in all_edges]
                    
    
    all_edges_lt = []
    [all_edges_lt.append(j.getLeadTime()) for j in all_edges]

    used_edges = dict(zip(all_edges, all_edges_lt))
    print(used_edges)
    
    return t, used_edges, all_cand_routes, last_arrival


def creates_template_schedule(graph, rhash_nodes, date, time_window): # to edit
    path = os.getcwd() + '\\network_schedule.xlsx'
    if not os.path.exists(path):
        writer = pd.ExcelWriter('network_schedule.xlsx')
        writer.save()
    wb = openpyxl.load_workbook('network_schedule.xlsx')
    if date.strftime('%d %b %Y ') in wb.sheetnames:
        print('\nThe Template is already generated.')
        return
    sheet = wb.create_sheet(title = date.strftime('%d %b %Y '))
    sheet['A1'] = 'Origin'
    sheet['B1'] = 'Destination'
    
    aux = list(string.ascii_uppercase)
    columns = [i+j for i in aux for j in aux]
    columns = aux[2:] + columns
    
    for i in columns[:time_window]:
        date += datetime.timedelta(days=1)
        sheet[i +'1'] = date.strftime('%a %d %b')
    row = 2
    for i in graph.edges:
        for j in graph.edges[i]:
            sheet['A' + str(row)] = rhash_nodes[int(j.getSource().getName())]
            sheet['B' + str(row)] = rhash_nodes[int(j.getDestination().getName())]
            row += 1
    wb.save('network_schedule.xlsx')
    print("\nIt is your turn to fill the sheet network 'schedule.xlsx'.")
    pass


def load_network_schedule(Sheetname, used_edges, rhash_nodes):
    answer = str()
    while answer != 'OK':
        answer = input("Introduce the availability of the routes: (write 'OK' when finish)\n")
    xl = pd.ExcelFile('network_schedule.xlsx')
    df = xl.parse(Sheetname) # parse the sheet
    reduced_schedule = pd.DataFrame()
    for i in used_edges:
        for j in range(df.shape[0]):
            if rhash_nodes[int(str(i.getSource()))] == df['Origin'][j] and \
                         rhash_nodes[int(str(i.getDestination()))] == df['Destination'][j]:
                reduced_schedule = pd.concat([reduced_schedule,df[j:j+1]])
    return reduced_schedule.sort_index()


def calc_rest_leadtime(graph, rest_path):
    acc = 0
    for i in range(len(rest_path)-1):
        acc += graph.getEdgeLT(rest_path[i],rest_path[i+1])
    return acc


def DFS_path_generator(t, graph, order, rest_path, vector, df, \
                       time_window, rhash_nodes, start_opt):
    """ cand_route is equivalent to rest_path"""
    if len(rest_path) == 1:
        vector.append('End')
        while len(vector) < time_window:
            vector.append('X')
        # I could save it in a DataFrame or in a transport structure
        t.addRoute(order, vector)
        return
    link = (rhash_nodes[int(str(rest_path[0]))], \
            rhash_nodes[int(str(rest_path[1]))])
    row = list(df[(df.Origin == link[0]) & \
             (df.Destination == link[1])].values[0][2:]) # returns a list of 1's and 0's
    for i in range(len(row)):
        # available connection, vector<i, and rest_time>=min_lt_to_do
        if (row[i] == True) and (len(vector) <= i) and \
        ((order.getDateEnd() - start_opt).days - len(vector) >= \
         calc_rest_leadtime(graph,rest_path)): # >= depends on start_opt
            new_vector = vector.copy()
            while (order.getDateStart() - start_opt).days > len(new_vector):
                new_vector.append('X') # order not ready to be sent
            while (len(new_vector) < i):
                new_vector.append('Wait')
            new_vector.append(str(rest_path[0])+'->'+str(rest_path[1]))
            rest_lt = graph.getEdgeLT(rest_path[0],rest_path[1])
            for j in range(rest_lt-1):
                new_vector.append('Transit')
            # recursive call with rest_path.pop[0] and new_vector
            if len(new_vector) <= ((order.getDateEnd() - start_opt).days): 
                new_rest_path = rest_path.copy()
                new_rest_path.pop(0)
                DFS_path_generator(t, graph, order, new_rest_path, new_vector, \
                                   df, time_window, rhash_nodes, start_opt) 


def populate_cons_matrix(t,g,all_cand_routes, reduced_schedule, time_window,\
                         rhash_nodes, start_opt):
    #DFS_path_generator(t,order,rest_path,vector=[],df,time_window,rhash_nodes)
    for i in range(len(all_cand_routes)):
        for j in range(len(all_cand_routes[i])): 
            DFS_path_generator(t, g, t.orders[i], all_cand_routes[i][j], [], \
                               reduced_schedule, time_window, rhash_nodes, start_opt)
    #consolidation_matrix = pd.DataFrame(t.routes)
    cons_matrix = []
    [cons_matrix.append([i.getName()] + j) for i in t.routes for j in t.routes[i]]
    label = []
    for i in reduced_schedule: label.append(i)
    label = ['Order'] + label[2:]
    df_cons_matrix = pd.DataFrame(cons_matrix, columns = label)
    return df_cons_matrix , cons_matrix

